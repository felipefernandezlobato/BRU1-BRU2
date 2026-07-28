#!/usr/bin/env python3
"""Import historical movement data from the COCINA - BRU Excel file into the database.

Reads the 'BRU1->BRU2' sheet from the Excel workbook, normalizes categories and
item names, and creates movements grouped by date.
"""
import os
import sys
from datetime import datetime
from collections import defaultdict

# Add parent dir to path so we can import app modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from app.database import SessionLocal
from app.models import Category, Item, Movement, MovementLine, Setting, CostHistory
from app.auth import hash_pin

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Path to the Excel workbook (most recent version)
EXCEL_PATH = os.environ.get(
    "IMPORT_EXCEL_PATH",
    "/Users/fernaf41/Downloads/COCINA - BRÜ  (4).xlsx",
)
SHEET_NAME = "BRU1->BRU2"

MARKUP_PCT = 30.0  # matches the DB setting

# ---------------------------------------------------------------------------
# Category mapping: raw spreadsheet name -> clean display name
# ---------------------------------------------------------------------------
CATEGORY_MAP = {
    "cafe matcha chai choc": "Cafe & Te",
    "tartas": "Reposteria",
    "milk": "Lacteos",
    "bowl": "Bowls",
    "sandwich": "Sandwiches",
    "bagel": "Bagels",
    "Drinks Beer": "Cerveza",
    "Drinks Soft drinks": "Refrescos",
    "Drinks Wine": "Vinos",
    "retail": "Retail",
    "tazas": "Tazas & Loza",
    "consumibles": "Consumibles",
    "no consumibles": "No Consumibles",
    "no consumible": "No Consumibles",
    "take away": "Take Away",
    "fruta": "Fruta",
    "Sirope": "Siropes",
    "varios": "Varios",
    "varios ": "Varios",
}

# ---------------------------------------------------------------------------
# Item name normalization
# ---------------------------------------------------------------------------
ITEM_NAME_MAP = {
    # Bagels
    "begels de salmon": "Bagel de salmon",
    "Begels de salmón": "Bagel de salmon",
    "begels de salmón": "Bagel de salmon",
    "Begel York": "Bagel York",
    "begels de atún": "Bagel de atun",
    # Baked goods
    "banana bread": "Banana bread",
    "Banana Bread": "Banana bread",
    "brownies": "Brownies",
    "croissant": "Croissant",
    "pain au chocolat": "Pain au chocolat",
    # Bowls
    "bowl yogurt": "Bowl yogurt",
    "Bowl yogurt": "Bowl yogurt",
    "Bowl chia": "Bowl chia",
    # Sandwiches
    "sándwich pollo": "Sandwich pollo",
    "sándwich atún": "Sandwich atun",
    # Milk
    "leche normal": "Leche entera",
    "leche Avena": "Leche de avena",
    "leche sin lactosa": "Leche sin lactosa",
    # Ensaladas (appear in tartas category in the sheet)
    "ensaladas cesar": "Ensalada cesar",
    "ensaldas mediterranea": "Ensalada mediterranea",
}

# Categories whose items are "produced" (made in BRU1 kitchen)
PRODUCED_CATEGORIES_RAW = {"tartas", "sandwich", "bagel", "bowl"}

# Individual items that are produced regardless of category
PRODUCED_ITEMS = {"Pure de fresa"}


def normalize_item_name(raw_name: str, raw_category: str) -> str:
    """Normalize an item name to a canonical form."""
    name = raw_name.strip()

    # Check direct mapping first
    if name in ITEM_NAME_MAP:
        return ITEM_NAME_MAP[name]

    # Strip "Drinks " prefix (redundant with category)
    if name.startswith("Drinks "):
        name = name[7:].strip()

    # Strip "take away " prefix (redundant with category)
    if name.startswith("take away "):
        name = name[10:].strip()

    # Normalize "cafe" -> "Cafe" in coffee item names
    if name.startswith("cafe "):
        name = "Cafe " + name[5:]

    # Strip "Sirope " prefix if category is already Siropes
    if raw_category.strip().lower() == "sirope" and name.startswith("Sirope "):
        name = name[7:].strip()
        # Capitalize first letter
        if name:
            name = name[0].upper() + name[1:]

    return name


def is_produced(raw_category: str, item_name: str) -> bool:
    """Determine if an item is produced (made in BRU1 kitchen)."""
    cat = raw_category.strip().lower()
    if cat in PRODUCED_CATEGORIES_RAW:
        return True
    if item_name in PRODUCED_ITEMS:
        return True
    return False


def parse_cost(valor) -> float:
    """Parse a cost value from the spreadsheet."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        s = valor.strip()
        if not s or s == "-":
            return 0.0
        # Remove "CHF " prefix
        s = s.replace("CHF", "").strip()
        # Remove Swiss thousand separator (apostrophe)
        s = s.replace("'", "")
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def parse_qty(qty) -> float | None:
    """Parse a quantity value. Returns None if invalid/skip."""
    if qty is None:
        return None
    if isinstance(qty, str):
        s = qty.strip()
        if not s or s == "-":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(qty, (int, float)):
        if qty == 0:
            return None
        return float(qty)
    return None


def extract_rows_from_excel(path: str, sheet_name: str) -> list[dict]:
    """Read and parse all valid data rows from the Excel sheet.

    Returns a list of dicts with keys:
        date (datetime.date), raw_category (str), raw_item (str),
        qty (float), cost_per_unit (float)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7, values_only=True):
        row_num, fecha, cat, item, qty, valor, total = row

        # Skip rows without fecha/cat/item
        if fecha is None and cat is None and item is None:
            continue

        # Skip rows without category (Personal entries, etc.)
        if cat is None or (isinstance(cat, str) and cat.strip() == ""):
            continue

        # Skip rows without item name
        if item is None or (isinstance(item, str) and item.strip() == ""):
            continue

        # Parse and validate qty
        parsed_qty = parse_qty(qty)
        if parsed_qty is None:
            continue

        # Parse date
        if isinstance(fecha, datetime):
            date_val = fecha.date()
        elif isinstance(fecha, str):
            # Shouldn't happen with data_only=True but handle it
            continue
        else:
            continue

        raw_category = cat.strip() if isinstance(cat, str) else str(cat)
        raw_item = item.strip() if isinstance(item, str) else str(item)
        cost = parse_cost(valor)

        rows.append({
            "date": date_val,
            "raw_category": raw_category,
            "raw_item": raw_item,
            "qty": parsed_qty,
            "cost_per_unit": cost,
        })

    wb.close()
    return rows


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at: {EXCEL_PATH}")
        print("Set IMPORT_EXCEL_PATH environment variable to the correct path.")
        sys.exit(1)

    print(f"Reading data from: {EXCEL_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    print()

    # -----------------------------------------------------------------------
    # 1. Parse raw data from Excel
    # -----------------------------------------------------------------------
    raw_rows = extract_rows_from_excel(EXCEL_PATH, SHEET_NAME)
    print(f"Parsed {len(raw_rows)} valid data rows from Excel")

    # -----------------------------------------------------------------------
    # 2. Normalize categories and items, deduplicate
    # -----------------------------------------------------------------------
    # Track unique categories and items
    category_names = set()         # clean category names
    # item_key -> {name, category, cost, is_produced, unit}
    items_dict: dict[tuple[str, str], dict] = {}
    # date -> list of (item_key, qty, cost_per_unit)
    movements_by_date: dict = defaultdict(list)

    for row in raw_rows:
        raw_cat = row["raw_category"]
        raw_item = row["raw_item"]

        # Map category
        clean_cat = CATEGORY_MAP.get(raw_cat, CATEGORY_MAP.get(raw_cat.strip(), raw_cat))
        category_names.add(clean_cat)

        # Normalize item name
        clean_item = normalize_item_name(raw_item, raw_cat)
        produced = is_produced(raw_cat, clean_item)

        item_key = (clean_cat, clean_item)

        # Update item info (last seen cost wins)
        cost = row["cost_per_unit"]
        if item_key not in items_dict:
            items_dict[item_key] = {
                "name": clean_item,
                "category": clean_cat,
                "cost": cost,
                "is_produced": produced,
                "unit": "ud",
            }
        else:
            # Update cost if new cost is non-zero, or keep existing
            if cost > 0:
                items_dict[item_key]["cost"] = cost

        # Add to movements grouped by date
        movements_by_date[row["date"]].append({
            "item_key": item_key,
            "qty": row["qty"],
            "cost_per_unit": cost,
        })

    print(f"Unique categories: {len(category_names)}")
    print(f"Unique items: {len(items_dict)}")
    print(f"Unique dates (movements): {len(movements_by_date)}")
    print()

    # -----------------------------------------------------------------------
    # 3. Write to database
    # -----------------------------------------------------------------------
    db = SessionLocal()
    try:
        # --- Clear existing data (keep users) ---
        print("Clearing existing data...")
        db.query(MovementLine).delete()
        db.query(Movement).delete()
        db.query(CostHistory).delete()
        db.query(Item).delete()
        db.query(Category).delete()
        db.query(Setting).delete()
        db.flush()
        print("  Cleared movement_lines, movements, cost_history, items, categories, settings")

        # --- Re-seed settings ---
        db.add(Setting(key="markup_pct", value="30"))
        db.add(Setting(key="escandallos_api_url", value="https://bru-escandallos-api.onrender.com"))
        db.flush()
        print("  Re-seeded settings")

        # --- Get admin user id ---
        from app.models import User
        admin = db.query(User).filter(User.role == "admin").first()
        if not admin:
            print("ERROR: No admin user found. Run the app first to seed the admin user.")
            sys.exit(1)
        admin_id = admin.id
        print(f"  Using admin user: {admin.name} (id={admin_id})")
        print()

        # --- Create categories ---
        print("Creating categories...")
        cat_db_map: dict[str, Category] = {}
        for i, cat_name in enumerate(sorted(category_names), start=1):
            cat = Category(name=cat_name, position=i, is_active=True)
            db.add(cat)
            db.flush()
            cat_db_map[cat_name] = cat
            print(f"  [{cat.id}] {cat_name}")
        print(f"  Created {len(cat_db_map)} categories")
        print()

        # --- Create items ---
        print("Creating items...")
        item_db_map: dict[tuple[str, str], Item] = {}
        for item_key, info in sorted(items_dict.items(), key=lambda x: (x[0][0], x[0][1])):
            cat_name, item_name = item_key
            cat_obj = cat_db_map[cat_name]
            item = Item(
                name=info["name"],
                category_id=cat_obj.id,
                unit=info["unit"],
                cost_per_unit=info["cost"],
                is_produced=info["is_produced"],
                is_active=True,
            )
            db.add(item)
            db.flush()
            item_db_map[item_key] = item
        print(f"  Created {len(item_db_map)} items")
        print()

        # --- Create movements (one per date) ---
        print("Creating movements...")
        total_lines = 0
        total_cost = 0.0
        sorted_dates = sorted(movements_by_date.keys())

        for mv_date in sorted_dates:
            lines_data = movements_by_date[mv_date]
            date_str = mv_date.strftime("%Y-%m-%d")

            movement = Movement(
                direction="BRU1_TO_BRU2",
                created_by=admin_id,
                notes=f"Historical import: {date_str}",
                movement_date=date_str,
            )
            db.add(movement)
            db.flush()

            for line_data in lines_data:
                item_key = line_data["item_key"]
                item_obj = item_db_map[item_key]
                cost = line_data["cost_per_unit"]
                qty = line_data["qty"]

                # Calculate transfer price (same logic as app)
                if item_obj.is_produced:
                    transfer_price = round(cost * (1 + MARKUP_PCT / 100), 4)
                    markup = MARKUP_PCT
                else:
                    transfer_price = cost
                    markup = 0.0

                ml = MovementLine(
                    movement_id=movement.id,
                    item_id=item_obj.id,
                    quantity=qty,
                    unit=item_obj.unit,
                    cost_per_unit_snapshot=cost,
                    markup_pct_snapshot=markup,
                    transfer_price_snapshot=transfer_price,
                )
                db.add(ml)
                total_lines += 1
                total_cost += transfer_price * qty

            db.flush()

        print(f"  Created {len(sorted_dates)} movements")
        print(f"  Created {total_lines} movement lines")
        print()

        # --- Commit ---
        db.commit()
        print("=" * 60)
        print("IMPORT COMPLETE")
        print("=" * 60)
        print(f"  Categories created:    {len(cat_db_map)}")
        print(f"  Items created:         {len(item_db_map)}")
        print(f"  Movements created:     {len(sorted_dates)}")
        print(f"  Movement lines:        {total_lines}")
        print(f"  Total cost:            CHF {total_cost:,.2f}")
        print(f"  Date range:            {sorted_dates[0]} to {sorted_dates[-1]}")

    except Exception as e:
        db.rollback()
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
